# xlsx_apply_mask.py
# 역할:
#  - 원본 XLSX와 마스킹 결과(.csv/.jsonl)를 받아 셀 값을 마스킹 값으로 치환
# 특징:
#  - openpyxl로 시트/서식 보존
#  - 검증(verify) 기본 동작:
#       * masked_text가 있을 때(= both/masked_only 출력): 셀의 현재값 == 원문(text)일 때만 교체
#       * masked_text가 없을 때(= replace 출력): 원문 없음 → 검증 자동 생략
#  - 옵션: --verify / --no-verify 로 강제 제어 가능, --force 는 검증 여부와 무관하게 무조건 치환
#  - 저장 경로 --out, 미지정 시 "원본명_masked.xlsx"
#
# 사용 예:
#   # replace 출력 사용(검증 자동 생략)
#   python xlsx_apply_mask.py "C:\data\원본.xlsx" "C:\data\원본.xlsx.masked_replace.csv"
#   # both/masked_only 출력 사용(검증 수행)
#   python xlsx_apply_mask.py "원본.xlsx" "원본.xlsx.masked_only.jsonl" --out "결과.xlsx"
#   # 어떤 경우든 무조건 치환
#   python xlsx_apply_mask.py "원본.xlsx" "원본.xlsx.masked_only.csv" --force

import argparse, csv, json, sys
from pathlib import Path
from typing import Dict, List, Tuple, Iterable, Optional
from openpyxl import load_workbook

# ---------- 입력 로딩 ----------
def read_pii_file(path: Path) -> List[Dict]:
    """
    지원: .csv / .jsonl
    필요한 필드:
      - source_type == "xlsx" 인 레코드만 적용
      - container(시트명), row(1-based), col(1-based)
      - masked_text(우선) 또는 text(대체; replace 출력은 text=마스킹값)
    """
    if path.suffix.lower() == ".csv":
        import pandas as pd
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
        return [dict(rec) for _, rec in df.iterrows()]
    elif path.suffix.lower() == ".jsonl":
        rows = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if s:
                    rows.append(json.loads(s))
        return rows
    else:
        raise SystemExit("지원 입력: .csv 또는 .jsonl")

def pick_masked_value(rec: Dict) -> Optional[str]:
    """masked_text 우선, 없으면 text(= replace 출력의 마스킹값). 비어있으면 None."""
    if "masked_text" in rec and str(rec["masked_text"]).strip() != "":
        return str(rec["masked_text"])
    if "text" in rec and str(rec["text"]).strip() != "":
        return str(rec["text"])
    return None

def same_source_file(rec: Dict, target_path: Path) -> bool:
    """source_path가 있으면 동일 파일인지 확인. 없으면 허용."""
    src = str(rec.get("source_path", "")).strip()
    if not src:
        return True
    try:
        return Path(src).resolve() == target_path.resolve()
    except Exception:
        return True  # 비교 실패 시 허용

# ---------- 적용 로직 ----------
def apply_masks_to_workbook(
    xlsx_path: Path,
    pii_rows: List[Dict],
    out_path: Path,
    force: bool = False,
    verify: Optional[bool] = None,  # None=자동: masked_text 유무에 따라
    dry_run: bool = False
) -> Tuple[int, int, int]:
    """
    반환: (총 후보 수, 적용 수, 스킵 수)
    """
    wb = load_workbook(filename=str(xlsx_path))
    total, applied, skipped = 0, 0, 0

    for rec in pii_rows:
        # xlsx에서 추출된 항목만 반영
        if str(rec.get("source_type", "")).lower() != "xlsx":
            continue
        if not same_source_file(rec, xlsx_path):
            continue

        sheet_name = str(rec.get("container", "")).strip()
        if not sheet_name:
            skipped += 1
            continue

        # 좌표(1-based)
        try:
            r = int(rec.get("row")) if rec.get("row") not in (None, "", "None") else None
            c = int(rec.get("col")) if rec.get("col") not in (None, "", "None") else None
        except Exception:
            skipped += 1
            continue
        if r is None or c is None or r < 1 or c < 1:
            skipped += 1
            continue

        masked_val = pick_masked_value(rec)
        if masked_val is None:
            skipped += 1
            continue

        # 검증 모드 결정
        has_masked_text = ("masked_text" in rec and str(rec["masked_text"]).strip() != "")
        # verify=None(자동)인 경우:
        #  - masked_text가 있으면 검증 수행(= 원문 존재)
        #  - masked_text가 없으면(= replace 출력) 검증 생략
        effective_verify = has_masked_text if (verify is None) else verify

        total += 1

        if sheet_name not in wb.sheetnames:
            skipped += 1
            continue
        ws = wb[sheet_name]
        cell = ws.cell(row=r, column=c)

        if not force and effective_verify:
            # 현재 셀 값과 원문(text)이 같을 때만 치환
            original_text = str(rec.get("text", ""))  # both/masked_only에서 text=원문
            cell_val = "" if cell.value is None else str(cell.value)
            if original_text and (cell_val != original_text):
                skipped += 1
                continue

        # 치환
        if not dry_run:
            cell.value = masked_val
        applied += 1

    # 저장
    if not dry_run:
        out_path = out_path.resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        print(f"[OK] writing: {out_path}")
        wb.save(str(out_path))

    return total, applied, skipped

# ---------- CLI ----------
def main():
    ap = argparse.ArgumentParser(description="Apply masked PII values back to original XLSX")
    ap.add_argument("xlsx", help="원본 XLSX 경로")
    ap.add_argument("masked_pii", help="마스킹 결과(.csv/.jsonl)")
    ap.add_argument("--out", default="", help="출력 XLSX 경로(기본: 원본명_masked.xlsx)")
    group = ap.add_mutually_exclusive_group()
    group.add_argument("--verify", dest="verify", action="store_true",
                       help="항상 원문(text) 검증 후 치환 (replace 출력에도 검증 강제)")
    group.add_argument("--no-verify", dest="verify", action="store_false",
                       help="항상 검증 생략 (both/masked_only 출력에도 검증 끔)")
    ap.set_defaults(verify=None)  # None=자동(설명 참조)
    ap.add_argument("--force", action="store_true", help="검증 여부와 무관하게 무조건 치환")
    ap.add_argument("--dry-run", action="store_true", help="저장하지 않고 통계만 출력")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"[ERR] 원본 XLSX 경로 없음: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    pii_path = Path(args.masked_pii)
    if not pii_path.exists():
        print(f"[ERR] 마스킹 결과 경로 없음: {pii_path}", file=sys.stderr)
        sys.exit(1)

    out_path = Path(args.out).resolve() if args.out else xlsx_path.with_name(xlsx_path.stem + "_masked.xlsx").resolve()

    # 로드
    pii_rows = read_pii_file(pii_path)

    # 적용
    total, applied, skipped = apply_masks_to_workbook(
        xlsx_path=xlsx_path,
        pii_rows=pii_rows,
        out_path=out_path,
        force=args.force,
        verify=args.verify,
        dry_run=args.dry_run
    )

    print(f"[DONE] candidates={total}, applied={applied}, skipped={skipped}")
    if not args.dry_run:
        print(f"[OUT] {out_path}")

if __name__ == "__main__":
    main()
